from openpyxl import load_workbook
from openpyxl.styles import PatternFill

filename = "2024加盟主問卷調查表_南區.xlsx"

TITLE_COLOR = PatternFill(start_color='D86DCD', end_color='D86DCD', fill_type='solid')
wb = load_workbook(filename)
ws = wb.active

branches = int(ws["B7"].value)

end = 8
while ws[f'A{end}'].value != "其他建議事項":
    end = end + 1
print(end)

for i in range(8, end, branches+1):
    ws.unmerge_cells(f"A{i}:A{i+branches}")
    
    block_value = ws[f"A{i}"].value
    ws[f"H{i}"].value = "point"
    for c in range(1, 9):
        ws.cell(row=i, column=c).fill = TITLE_COLOR
    
    for x in range(i+1, i+branches+1):
        ws[f"A{x}"].value = block_value
        points = int(ws[f"D{x}"].value) * 1 + int(ws[f"E{x}"].value) * 2 + int(ws[f"F{x}"].value * 3) + int(ws[f"G{x}"].value * 4)
        ws[f"H{x}"].value = points
        
wb.save("./2024加盟主問卷調查表_南區new.xlsx")